#strings
print("200 is a great number")
# numbers
print(2)
#Boolean
# print(-2>-5)

print("20 days are xxx minites")
print(20*24*60)
#String concatination (f stands for format and only works for python3)
print("20 days are "+ str(50) + " minutes")
print(f"20 days are {50} minutes")

# Variables
calculation_to_units = 24 * 60 * 60 # no need to specify the datatype, python automatically picks the data type unlike other languages
name_of_unit = "seconds"
print(f"20 days are {20 * calculation_to_units} {name_of_unit}")
print(f"20 days are {35 * calculation_to_units} {name_of_unit}")
print(f"20 days are {50 * calculation_to_units} {name_of_unit}")
print(f"20 days are {110 * calculation_to_units} {name_of_unit}")
# be careful abour reserved words, cannot be used as vars

# functions (define function)"block of code that does something to avoid using the repeating logic"
# when defining function it expects 2 blank line before function
calculation_to_units = 24 * 60 * 60 # no need to specify the datatype, python automatically picks the data type unlike other languages
name_of_unit = "seconds"


def days_to_units(num_of_days, custome_message):
    print(f"{num_of_days} days are {num_of_days * calculation_to_units} {name_of_unit}")
    print(custome_message)


days_to_units(20, "Awsome!")  # calling the function
days_to_units(35, "Looks Good!")

# Scope  (A variable is only available form inside the region it is created)

def scope_check(num_of_days):
    my_var = "variable inside functions"
    print(name_of_unit)
    print(num_of_days)
    print(my_var)
scope_check(20) # error if num_of_days not defined inside

# User input ; input function used
calculation_to_units = 24 * 60 * 60 # no need to specify the datatype, python automatically picks the data type unlike other languages
name_of_unit = "seconds"

def days_to_units(num_of_days):
    print(f"{num_of_days} days are {num_of_days * calculation_to_units} {name_of_unit}")

user_input = input("Hey user, enter a number of days and I will conver it to seconds!\n")
print(user_input)

# Function with Return Values
calculation_to_units = 24 * 60 * 60 # no need to specify the datatype, python automatically picks the data type unlike other languages
name_of_unit = "seconds"

def days_to_units(num_of_days):
    return(f"{num_of_days} days are {num_of_days * calculation_to_units} {name_of_unit}")

my_var = days_to_units(35)
print(my_var)

# Function with Return Values
# User input ; input function used
calculation_to_units = 24 * 60 * 60 # no need to specify the datatype, python automatically picks the data type unlike other languages
name_of_unit = "seconds"

def days_to_units(num_of_days):
    return(f"{num_of_days} days are {num_of_days * calculation_to_units} {name_of_unit}")

# input values we give in input is always treated as a string not number, so we need specify int
user_input = input("Hey user, enter a number of days and I will conver it to seconds!\n")
user_input_number = int(user_input)

calculated_values = days_to_units(user_input_number)
print(calculated_values)

# Restricting users to only add spefic  input
# Validate User input
# Using "Conditionals" if and else
calculation_to_units = 24 * 60 * 60 # no need to specify the datatype, python automatically picks the data type unlike other languages
name_of_unit = "seconds"

def days_to_units(num_of_days):
    conditional_check = num_of_days > 0
    print(type(conditional_check))
    if num_of_days > 0:    # conditional
        return f"{num_of_days} days are {num_of_days * calculation_to_units} {name_of_unit}"
    else:
        return "you entered negative value, so no convertion for you"


# input values we give in input is always treated as a string not number, so we need specify int
user_input = input("Hey user, enter a number of days and I will conver it to seconds!\n")

calculated_values = days_to_units(int(user_input))
print(calculated_values)

# Two equal signs and conditionals
calculation_to_units = 24 * 60 * 60 # no need to specify the datatype, python automatically picks the data type unlike other languages
name_of_unit = "seconds"

def days_to_units(num_of_days):
    conditional_check = num_of_days > 0
    print(type(conditional_check))
    if num_of_days > 0:    # conditional
        return f"{num_of_days} days are {num_of_days * calculation_to_units} {name_of_unit}"
    elif num_of_days == 0:   # two equal signs when checking if equals
        return "you entered 0, please enter valid positive value"
    else:
        return "you entered negative value, so no convertion for you"


# input values we give in input is always treated as a string not number, so we need specify int
user_input = input("Hey user, enter a number of days and I will conver it to seconds!\n")

calculated_values = days_to_units(int(user_input))
print(calculated_values)

# validate users input
calculation_to_units = 24 * 60 * 60 # no need to specify the datatype, python automatically picks the data type unlike other languages
name_of_unit = "seconds"

def days_to_units(num_of_days):
    conditional_check = num_of_days > 0
    print(type(conditional_check))
    if num_of_days > 0:    # conditional
        return f"{num_of_days} days are {num_of_days * calculation_to_units} {name_of_unit}"
    elif num_of_days == 0:   # two equal signs when checking if equals
        return "you entered 0, please enter valid positive value"
    else:
        return "you entered negative value, so no convertion for you"


# input values we give in input is always treated as a string not number, so we need specify int
user_input = input("Hey user, enter a number of days and I will conver it to seconds!\n")
# validation of user input
if user_input.isdigit():
    user_input_number = int(user_input)
    calculated_values = days_to_units(user_input_number)
    print(calculated_values)
else:
    print("Your input is not a number, Don't ruin my program")

# Cleanup
# validate users input
calculation_to_units = 24 * 60 * 60 # no need to specify the datatype, python automatically picks the data type unlike other languages
name_of_unit = "seconds"

def days_to_units(num_of_days):
    if num_of_days > 0:    # conditional
        return f"{num_of_days} days are {num_of_days * calculation_to_units} {name_of_unit}"
    elif num_of_days == 0:   # two equal signs when checking if equals
        return "you entered 0, please enter valid positive value"

def validate_and_execute():
    if user_input.isdigit():
        user_input_number = int(user_input)
        calculated_values = days_to_units(user_input_number)
        print(calculated_values)
    else:
        print("Your input is not a number, don't ruin my program")


# input values we give in input is always treated as a string not number, so we need specify int
user_input = input("Hey user, enter a number of days and I will conver it to seconds!\n")
validate_and_execute()

# Nested if else
# validate users input
calculation_to_units = 24 * 60 * 60 # no need to specify the datatype, python automatically picks the data type unlike other languages
name_of_unit = "seconds"

def days_to_units(num_of_days):
    return f"{num_of_days} days are {num_of_days * calculation_to_units} {name_of_unit}"

def validate_and_execute():
    if user_input.isdigit():
        user_input_number = int(user_input)
        if user_input_number > 0:  # conditional
            calculated_values = days_to_units(user_input_number)
            print(calculated_values)
        elif user_input_number == 0:  # two equal signs when checking if equals
            print("you entered 0, please enter valid positive value")
    else:
        print("Your input is not a number, don't ruin my program")


# input values we give in input is always treated as a string not number, so we need specify int
user_input = input("Hey user, enter a number of days and I will conver it to seconds!\n")
validate_and_execute()

# try except; Also known as try catch in other programing laguange
calculation_to_units = 24 * 60 * 60 # no need to specify the datatype, python automatically picks the data type unlike other languages
name_of_unit = "seconds"

def days_to_units(num_of_days):
    return f"{num_of_days} days are {num_of_days * calculation_to_units} {name_of_unit}"

def validate_and_execute():
    try:

        user_input_number = int(user_input)
        if user_input_number > 0:  # conditional
            calculated_values = days_to_units(user_input_number)
            print(calculated_values)
        elif user_input_number == 0:  # two equal signs when checking if equals
            print("you entered 0, please enter valid positive value")
        else:
            print("you entered a negative number, no conversion for you")
    except ValueError:
        print("Your input is not a number, don't ruin my program")


# input values we give in input is always treated as a string not number, so we need specify int
user_input = input("Hey user, enter a number of days and I will conver it to seconds!\n")
validate_and_execute()

# while loop; execute a set of statements as long as a condition is true
#  Let user exit the program
calculation_to_units = 24 * 60 * 60 # no need to specify the datatype, python automatically picks the data type unlike other languages
name_of_unit = "seconds"

def days_to_units(num_of_days):
    return f"{num_of_days} days are {num_of_days * calculation_to_units} {name_of_unit}"

def validate_and_execute():
    try:

        user_input_number = int(user_input)
        if user_input_number > 0:  # conditional
            calculated_values = days_to_units(user_input_number)
            print(calculated_values)
        elif user_input_number == 0:  # two equal signs when checking if equals
            print("you entered 0, please enter valid positive value")
        else:
            print("you entered a negative number, no conversion for you")
    except ValueError:
        print("Your input is not a number, don't ruin my program")

user_input = ""    # so while loop doesn't through error
while user_input != "exit":
    user_input = input("Hey user, enter a number of days and I will conver it to seconds!\n")
    validate_and_execute()

# List ; to store multiple items in a single variable
#  A list can contain different data types
# "For" loop = used for iterating over a sequence; so we can exercise something for each item in a list
calculation_to_units = 24 * 60 * 60 # no need to specify the datatype, python automatically picks the data type unlike other languages
name_of_unit = "seconds"

def days_to_units(num_of_days):
    return f"{num_of_days} days are {num_of_days * calculation_to_units} {name_of_unit}"

def validate_and_execute():
    try:
        user_input_number = int(num_of_days_element)
        if user_input_number > 0:  # conditional
            calculated_values = days_to_units(user_input_number)
            print(calculated_values)
        elif user_input_number == 0:  # two equal signs when checking if equals
            print("you entered 0, please enter valid positive value")
        else:
            print("you entered a negative number, no conversion for you")
    except ValueError:
        print("Your input is not a number, don't ruin my program")

user_input = ""    # so while loop doesn't through error
while user_input != "exit":
    user_input = input("Hey user, enter a number of days as a comma separated list and I will conver it to seconds!\n")
    print(type(user_input.split(",")))
    print(user_input.split(","))
    for num_of_days_element in user_input.split(","):      # split function will take the string "10,20,50,100" and convert into list [10, 20, 50, 100]
        validate_and_execute()     # to excute just add 10 20 30 with space or commans if we add it in split funtion

# List examples and append to add values to the vairable
my_list =["January", "February", "March"]
print(my_list[2])
my_list.append("Apirl")   # add values to the list
print(my_list)

# Comments - used to explain code
""" multi line comments
test scenario varies"""

# Sets- another built-in data type of python
# as with lists, used to store multiple items of data
# Difference from list, Duplicate values not allowed
calculation_to_units = 24 * 60 * 60 # no need to specify the datatype, python automatically picks the data type unlike other languages
name_of_unit = "seconds"

def days_to_units(num_of_days):
    return f"{num_of_days} days are {num_of_days * calculation_to_units} {name_of_unit}"

def validate_and_execute():
    try:
        user_input_number = int(num_of_days_element)
        if user_input_number > 0:  # conditional
            calculated_values = days_to_units(user_input_number)
            print(calculated_values)
        elif user_input_number == 0:  # two equal signs when checking if equals
            print("you entered 0, please enter valid positive value")
        else:
            print("you entered a negative number, no conversion for you")
    except ValueError:
        print("Your input is not a number, don't ruin my program")

user_input = ""    # so while loop doesn't through error
while user_input != "exit":
    user_input = input("Hey user, enter a number of days as a comma separated list and I will conver it to seconds!\n")
    list_of_days = user_input.split(", ")
    print(list_of_days)
    print(set(list_of_days))
    print(type(list_of_days))
    print(type(set(list_of_days)))

    for num_of_days_element in set(list_of_days):      # split function will take the string "10,20,50,100" and convert into list [10, 20, 50, 100]
        validate_and_execute()     # to excute just add 10 20 30 with space or commans if we add it in split funtion

# ***
# Built-In Functions : we've used so far - print(), input(), set(), int()
# "2 , 3".split() -> functions called directly on the values
# ***

# ***
# to view the list
my_list = ["20", "30", "100"]
print(my_list[1])

# to print the dictionary
my_dictionary = {"days": 20, "unit": "hours"}
print(my_dictionary["days "])
# ***

# Dictionary: used to store values in key:values pairs; Is a collection, which does not allow duplicate values
def days_to_units(num_of_days, conversion_unit):
    if conversion_unit == "hours":
        return f"{num_of_days} days are {num_of_days * 24} hours"
    elif conversion_unit == "minutes":
        return f"{num_of_days} days are {num_of_days * 24 * 60} minutes"
    else:
        return "unsupported unit"
def validate_and_execute():
    try:
        user_input_number = int(days_and_unit_dictionary["days"])
        if user_input_number > 0:  # conditional
            calculated_values = days_to_units(user_input_number, days_and_unit_dictionary["unit"])
            print(calculated_values)
        elif user_input_number == 0:  # two equal signs when checking if equals
            print("you entered 0, please enter valid positive value")
        else:
            print("you entered a negative number, no conversion for you")
    except ValueError:
        print("Your input is not a number, don't ruin my program")

user_input = ""    # so while loop doesn't through error
while user_input != "exit":
    user_input = input("Hey user, enter a number of days and conversion unit!\n")
    days_and_unit = user_input.split(":") # split function will take the string "10,20,50,100" and convert into list [10, 20, 50, 100]
    print(days_and_unit)
    days_and_unit_dictionary = {"days": days_and_unit[0], "unit": days_and_unit[1]}  #syntax of dictionary
    print(days_and_unit_dictionary)
    print(type(days_and_unit_dictionary))
    validate_and_execute()

#*** learn type of data types we learned so far
message = "enter some value" # string data type
days = 20 # int
price = 9.99 #float
valid_number = True # boolean
list_of_days = [20, 40 ,60, 20] #list string iny
set_of_days = {20, 45, 100} #set
days_to_units = {"days": 10 "unit": "hours"} # dictionary data type
#***

# modules: logically organise your python code;
# modules is just a python file that contains functions or variables athat we can use in another python file

import helper # we import entire module and refer in the funtion below

user_input = ""    # so while loop doesn't through error
while user_input != "exit":
    user_input = input("Hey user, enter a number of days and conversion unit!\n")
      days_and_unit = user_input.split(":") # split function will take the string "10,20,50,100" and convert into list [10, 20, 50, 100]
    days_and_unit_dictionary = {"days": days_and_unit[0], "unit": days_and_unit[1]}  #syntax of dictionary
    helper.validate_and_execute(days_and_unit_dictionary)

# *** if we only want to import specific funtion
from helper import validate_and_execute
user_input = ""    # so while loop doesn't through error
while user_input != "exit":
    user_input = input("Hey user, enter a number of days and conversion unit!\n")
      days_and_unit = user_input.split(":") # split function will take the string "10,20,50,100" and convert into list [10, 20, 50, 100]
    days_and_unit_dictionary = {"days": days_and_unit[0], "unit": days_and_unit[1]}  #syntax of dictionary
    validate_and_execute(days_and_unit_dictionary)

from helper import validate_and_execute, user_input_message # importing function and var
# from helper import * # to import all
# import helper as h # we can rename the module and access it as h
# the difference of using from helper import * instead of just module is that we don't need to user helper.function/var all the time like we use module helper
user_input = ""    # so while loop doesn't through error
while user_input != "exit":
    user_input = input(user_input_message)
      days_and_unit = user_input.split(":") # split function will take the string "10,20,50,100" and convert into list [10, 20, 50, 100]
    days_and_unit_dictionary = {"days": days_and_unit[0], "unit": days_and_unit[1]}  #syntax of dictionary
    validate_and_execute(days_and_unit_dictionary)

# built-in python modules
""" many existing modules
import math
from datetime import datetime, timezone
"""
import os  # operating system module
print(os.name) #name of the OS

import logging # if we want to use loggin in our application
logger = logging.getLogger("MAIN")
logger.error("Error happened in the app")

import datetime

user_input = input("enter your goal with a deadline separated by colon\n") # user input is always interpreted as string
input_list = user_input.split(":")

goal = input_list[0]
deadline = input_list[1]

print(datetime.datetime.strptime(deadline, "%d.%m.%Y")) #can be %d/%m/%y ; capital Y for full date small y for last 2 digit
print(type(datetime.datetime.strptime(deadline, "%d.%m.%Y")))
print(input_list)

# ***

# built-in vs Third-Party

'''we can find all third party modules like for django in pypi.org for p
python packages download, from there we can download a package

*Diff between package and module
Module - is just a python file that has python functions and variable
package = collection of modules, multiple modules file; each module might have more that
one py file
like logging module = __init__.py, config.py and handlers.py file in one module'''

'''******pip*****=python package manager
to install the packages onto our device using "pip"

pip is included inside the python

eg. pip install django
pip uninstall django'''

'''************'''

''''Automation with Python'''
''' Python has several built-in functions for handling files in general
io module -> create, read, write
python package to work with spreadsheets specifically = openpyxl
'''
'''
module = .py file
package = collection of init.py file
library = collection of packages'''

import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx") #load the xml file
product_list = inv_file["Sheet1"]   #name of the spread sheet

#Task 1 = calculate how many products per suppliers and list that product respective to suppliers
products_per_supplier = {}
total_value_per_supplier = {}
products_under_10_inv = {}
#Now we have through each row in the sheet
# Execute same logic on each item, as many times as number of products
# starting from 2 means starting from row number 2, skips the head
# range function is not exclusive, so we have to add 1 to productlist, range becomes 76, then 75 rows gets executed
for product_row in range(2, product_list.max_row + 1): # range creates a sequence of numbers, starting from 0 by default range75=[0,1,2..74];
    supplier_name = product_list.cell(product_row, 4).value # supplier name
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5) # value not needed as, as we need cell info
   # calculation for number of products per supplier
    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier.get(supplier_name) # instead of brackets we cal also use get
        products_per_supplier[supplier_name] = current_num_products + 1
    else:
        print("Adding a new supplier")
        products_per_supplier[supplier_name] = 1
    # calculation total value of inventory per supplier
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # logic products with inventory less than 10
    if inventory < 10 :
        products_under_10_inv[int(product_num)] = int(inventory)

    # add value for total inventory price
    inventory_price.value = inventory * price

print(products_under_10_inv)
print(products_per_supplier)
print(total_value_per_supplier)

inv_file.save("inventory_with_total_value.xlsx")

'''OBJECTS & CLASSES
Object-oriented programming '''

'''
The blueprint of the user is called class
like: user attributes: email, Name, password, job tittle
      user behaviour: change_pwd, change_job
Class will define: what information user has; what actions user can perform

The implementation of the blueprint is called object
Object then will have actual information of the person'''

# Create class user
import user   #using the file name, not class name

app_user_one = user.User("as@as.com", "Arsh Singh", "pwd1", "DevOps Engineer")
app_user_one.get_user_info()

app_user_two = user.User("ba@aba.com", "James Bond", "pwd1", "DevOps Engineer")
app_user_two.get_user_info()

# rewrite above using from
# Create class user
from user import User  #using the file name, not class name

app_user_one = User("as@as.com", "Arsh Singh", "pwd1", "DevOps Engineer")
app_user_one.get_user_info()

app_user_two = User("ba@aba.com", "James Bond", "pwd1", "DevOps Engineer")
app_user_two.get_user_info()

'''*********'''
from user import User  #using the file name, not class name
from post import Post
app_user_one = User("as@as.com", "Arsh Singh", "pwd1", "DevOps Engineer")
app_user_one.get_user_info()

app_user_two = User("ba@aba.com", "James Bond", "pwd1", "DevOps Engineer")
app_user_two.get_user_info()

new_post = Post("on a secret mission today", app_user_two.name)
new_post.get_post_info()

'''Object orinted programming
means when creating objects and classes to create those blueprints
**FAct
In python everything is an object = string, integer, float, set, dictionary
'''

'''API Request to Gitlab'''

#Python---->http request--->gitlab
#Python <----http response--gitlab

# We have a package in python; to make external request
# pip install requests ; Requests is http library
# we can also make request to change something on remote server

import requests

response = requests.get("https://gitlab.com/api/v4/users/nanuchi/projects")
my_projects = response.json()

for project in my_projects:
    print(f"Project Name: {project['name']}\n Project Url: {project['web_url']}\n") # because the json is converted to dictionary as type

# print(response.text)
# print(type(response.text))   # string data type
#
# print(response.json()) #request json() function converts data from json into python data type
# print(type(response.json())) # becomes list data type
# print(response.json()[0])

